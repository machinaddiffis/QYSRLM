import numpy as np
from openpyxl import load_workbook
from scipy import interpolate
import math

class Mcs_Sinr_p():
    def __init__(self,len_MCS=29,len_SINR=37):
        self.len_MCS = len_MCS
        self.len_SINR = len_SINR
        self.ori_P = np.zeros((len_SINR, len_MCS))
        self.SINR_class = []
        self.MCS_class = [i for i in range(29)]
        self.Function_list = []
    def load_table(self):
        workbook = load_workbook(filename='MCS_table.xlsx')
        # 选择工作表
        sheet = workbook['Sheet1']
        # 读取单元格内容
        cell_value = sheet['A1'].value
        # 读取整个工作表数据
        i = 0
        for row in sheet.iter_rows(min_row=1, max_col=30, values_only=True):
            if row[0] != 'SINR':
                self.SINR_class.append(row[0])
                self.ori_P[:][i] = row[1:]
                i += 1

    def create_inte(self):
        self.ori_P[np.isnan(self.ori_P)] = 0  # Y
        self.ori_P = np.array(self.ori_P)

        for j in range(len(self.MCS_class)):
            # 补0插值
            nums = self.len_SINR
            f = interpolate.interp1d(x=self.SINR_class[:nums], y=self.ori_P[:, j][:nums], kind="quadratic", bounds_error=False)
            self.Function_list.append(f)

    def Error_cal(self,pre_MCS,req_SINR,cls=29,probs=False):

        np.random.seed(42)
        UEs = pre_MCS.shape[0]
        error_p=[]
        for i in range(len(pre_MCS)):
            if pre_MCS[i] <= -1 or math.isnan( req_SINR[i]):
                p = 1
            else:
                if pre_MCS[i] >= 29:
                    pre_MCS[i] = 28

                p=self.Function_list[int(pre_MCS[i])](req_SINR[i])

            if np.isnan(p):
                # print("check!","MCS:",pre_MCS[i],"SINR:",req_SINR[i],"Probs:",p)
                #changed
                p=1
            elif p > 1:
                p = 1
            elif p < 0:
                p = 0

            error_p.append(p)

        event = (np.random.binomial(1, error_p, size=UEs))
        if probs==True:
            return error_p
        else:
            return event


    def Error_cal_single(self,pre_MCS,req_SINR,cls=29,probs=False):

        error_p=[]

        if req_SINR<=-1000:
            p=0
        if pre_MCS==-1:
            p=1
        else:
            p=self.Function_list[int(pre_MCS)](req_SINR)

        if np.isnan(p):
            # print("check!","MCS:",pre_MCS[i],"SINR:",req_SINR[i],"Probs:",p)
            #changed
            p=1
        if p<0:
            p=0
        if p>1:
            p=1
        error_p.append(p)
        event = (np.random.binomial(1, error_p[0], size=1))
        if probs==True:
            return error_p
        else:
            return event

def mcs_decrease(ref_mcs,de):

    ref_mcs = ref_mcs + de

    for i in range(ref_mcs.shape[1]):
        if ref_mcs[0, i] < -1:
            ref_mcs[0, i] = -1
        # if ref_mcs[0, i] == -1:
        #     ref_mcs[0, i] = 0
        if ref_mcs[0, i] > 28:
            ref_mcs[0, i] = 28
    return ref_mcs





if __name__ == "__main__":
    A=Mcs_Sinr_p()
    A.load_table()
    A.create_inte()
    SINR=np.random.randint(-8,28,size=30)
    MCS=np.random.randint(0,28,size=30)

    print(A.Error_cal(pre_MCS=MCS,req_SINR=SINR))

