import math
import random
import numpy as np
from openpyxl import load_workbook
from scipy import interpolate
import pickle
import torch
import torch.nn.functional as F
random.seed(0)
np.random.seed(0)

MCS_TABLE = [   0.1523, 0.2344, 0.377,  0.6016, 0.877,  1.1758, 1.4766, 1.6953, 1.9141, 2.1602,
                2.4063, 2.5703, 2.7305, 3.0293, 3.3223, 3.6094, 3.9023, 4.2129, 4.5234, 4.8164,
                5.1152, 5.332,  5.5547, 5.8906, 6.2266, 6.5703, 6.9141, 7.1602, 7.4063]

# SINR_TABLE = [0.31126, 0.09484, 0.13312, 0.14236, 0.15482, 0.12404, 0.13868, 0.1494, 0.14626, 0.17662,
#             0.17454, 0.1531, 0.12472, 0.13966, 0.1322, 0.13974, 0.13216, 0.12616, 0.1426, 0.15916,
#             0.16594, 0.18384, 0.25542, 0.29776, 0.11504, 0.14404, 0.289, 0.11964, 0.24792, 0.10134,
#             0.29456, 0.13976, 0.0438, 0.01042, 0.00206, 0.0003, 0.00018]
SINR_TABLE = [0, 0, 1, 2, 3, 4, 5, 6, 7, 8,
                9, 11, 12, 13, 14, 15, 16, 18, 19, 20,
                21, 22, 23, 24, 24, 25, 26, 26, 27, 27,
                28, 28, 28, 28, 28, 28, 28]
def random_points(dmin, dmax, N):
    """
    生成 N 个随机点，这些点到原点 (0,0) 的距离在 [dmin, dmax] 之间。

    :param dmin: 最小距离
    :param dmax: 最大距离
    :param N: 需要生成的点的数量
    :return: 一个包含 N 个点 (x, y) 坐标的列表
    """
    points = []
    for _ in range(N):
        r = random.uniform(dmin, dmax)  # 生成随机半径
        theta = random.uniform(0, 2 * math.pi)  # 生成随机角度
        x = r * math.cos(theta)
        y = r * math.sin(theta)
        points.append((x, y))
        # print(math.sqrt(x*x+y*y))
    return points

def beta_N(SP_N,UE_point,PL_0,d_0,alpha,VIP_TUNE=None):
    mu=1000
    b_N=[]
    if VIP_TUNE is None:
        for n in range(len(SP_N)):
            x,y=UE_point[n]
            dn=math.sqrt(x*x+y*y)
            PL_n=PL_0+SP_N[n]+10*alpha*math.log10((dn/d_0))
            b_N.append(math.pow(10,(-PL_n/10)))
        return b_N
    else:
        for n in range(len(SP_N)):
            x, y = UE_point[n]
            if n in VIP_TUNE:
                dn = math.sqrt(x * x + y * y)*1.5*mu
            else:
                dn = math.sqrt(x * x + y * y)*mu
            PL_n = PL_0 + SP_N[n] + 10 * alpha * math.log10((dn / d_0))
            b_N.append(math.pow(10, (-PL_n / 10)))
            # if n in VIP_TUNE:
            #
            #     print(math.pow(10, (-PL_n / 10)))
            #     print(-PL_n)
            #
            #     cs=math.sqrt(x * x + y * y)
            #     PL_n = PL_0 + SP_N[n] + 10 * alpha * math.log10((cs / d_0))
            #     print(math.pow(10, (-PL_n / 10)))
            #     print(-PL_n)
            #     quit()
        return b_N

    pass


def sample_vector(gamma, N,rest_list=None,time=None):
    """
    以系数 gamma 进行采样，返回为 1 的索引。

    :param gamma: 采样系数，表示每个元素成为 1 的概率
    :param N: 向量长度
    :return: 所有值为 1 的索引列表
    """
    if time is not None:
        random.seed(time)
        np.random.seed(time)
    if rest_list is None:
        vector = np.random.choice([0, 1], size=N, p=[1 - gamma, gamma])
        dis=list(np.where(vector == 1)[0])
    else:
        RE=1
        while RE>0:
            vector = np.random.choice([0, 1], size=len(rest_list), p=[1 - gamma, gamma])
            numpy_list=np.array(rest_list)

            right=numpy_list*vector
            # print(rest_list)
            print("采样结果:",np.where(right >0))
            dis=list(right[np.where(right >0)[0]])
            if len(right[np.where(right >0)[0]])>0:
                RE=-1

            # for i in range(3):
            #     vector = np.random.choice([0, 1], size=100, p=[1 - gamma, gamma])
            #     numpy_list = np.array(rest_list)
            #     right = numpy_list
            #     print("采样:", np.where(vector > 0))
            # quit()
    return dis


def remove_random_elements(vector, k):
    """
    从向量中随机删除 k 个元素。

    :param vector: 输入向量 (列表)
    :param k: 要删除的元素个数
    :return: 删除 k 个元素后的新向量
    """
    if k > len(vector):
        raise ValueError("k 不能大于向量的长度")
    indices_to_remove = set(random.sample(range(len(vector)), k))
    return [v for i, v in enumerate(vector) if i not in indices_to_remove]

def common_elements(A, B):
    # 使用集合求交集
    common=[item for item in A if item in B]
    # 如果交集为空，则返回None
    return common if common else None

def remove_elements(A, B):
    # 利用列表推导式过滤掉在B中的元素
    return [item for item in A if item not in B]


def generate_rayleigh_matrix(M, Tx,s=1):
    # 生成 n×m 矩阵的瑞利分布幅值
    r = np.random.rayleigh(scale=s, size=(M, Tx))
    # 生成 n×m 矩阵的均匀分布相位
    phi = np.random.uniform(0, 2 * np.pi, size=(M, Tx))
    # 构造复数矩阵
    matrix = r * np.exp(1j * phi)
    return matrix


def compute_zf_precoder(H):
    """
    计算零迫使（ZF）预编码矩阵
    参数:
        H: 信道矩阵，形状为 (num_RBG, num_users, N_tx, 1)
           假设每个用户只有一个接收天线
    返回:
        V: 零迫使预编码矩阵，形状为 (num_RBG, num_users, N_tx, 1)
    """
    print(H.shape,"shape")
    N,M, Tx = H.shape

    # V = np.zeros((num_RBG, num_users, N_tx, 1), dtype=H.dtype)
    # for i in range(num_RBG):
    #     # 对于第i个RBG，将所有用户的信道拼成一个 (num_users x N_tx) 的矩阵
    #     H_agg = H[i, :, :, 0]  # 形状: (num_users, N_tx)
    #     # 计算伪逆：当信道矩阵可能不满秩或数值不稳定时，推荐使用np.linalg.pinv
    #     # 这里可以写成：
    #     V_agg = np.linalg.pinv(H_agg)
    #     # V_agg 的形状为 (N_tx, num_users)
    #     # 将每个用户的预编码向量放回对应的位置
    #     for u in range(num_users):
    #         V[i, u, :, 0] = V_agg[:, u]
    #
    # return V
    V = np.zeros((N, Tx, M), dtype=complex)

    for i in range(N):
        Hi = H[i]  # 对应第i个RBG的信道矩阵，形状 (M, Tx)
        # 计算 H_i H_i^H，形状为 (M, M)
        Hi_H = Hi @ Hi.conj().T

        # 判断是否满秩，若满秩则直接求逆，否则使用伪逆
        if np.linalg.matrix_rank(Hi_H) == M:
            inv_term = np.linalg.inv(Hi_H)
        else:
            inv_term = np.linalg.pinv(Hi_H)

        # ZF预编码矩阵 V_i = H_i^H * inv(H_i * H_i^H)
        V[i] = Hi.conj().T @ inv_term

        # （可选）对预编码向量进行归一化（例如按列归一化）：
        # for j in range(M):
        #     norm_factor = np.linalg.norm(V[i, :, j])
        #     if norm_factor > 0:
        #         V[i, :, j] /= norm_factor

    return V

def compute_SINR(X, H, V, noise_power=1e-6):
    """
    计算每个用户在各资源块上的SINR

    参数:
    -----------
    X : np.array, shape (M, N)
        用户-资源块分配矩阵（0或1），X[m,n]==1 表示用户 m 被分配了资源块 n。
    H : np.array, shape (M, N, Tx)
        信道矩阵，描述每个用户在各资源块、各发射天线上的信道响应（可以为复数）。
    V : np.array, shape (M, N, Tx)
        权重/预编码矩阵，与信道矩阵对应（可为复数）。
    noise_power : float, 可选
        噪声功率，默认为 1e-3。

    返回:
    --------
    SINR : np.array, shape (M, N)
        每个用户在对应资源块上的SINR值，对于未分配的资源块，SINR值保持为0。
    """
    M, N, Tx = H.shape
    SINR = np.zeros((M, N))

    for m in range(M):
        for n in range(N):
            # 若用户 m 在资源块 n 上被分配
            if X[m, n] != 0:
                # 期望信号：用户 m 在 n 资源块上的内积
                desired = np.dot(H[m, n, :], V[m, n, :])
                signal_power = X[m, n]*np.abs(desired) ** 2

                interference_power = 0.0
                # 对于同一资源块 n 上分配给其他用户的信号，计算干扰
                for k in range(M):
                    if k != m and X[k, n] != 0:
                        interfering = np.dot(H[m, n, :], V[k, n, :])
                        interference_power += X[k, n]*np.abs(interfering) ** 2
                SINR[m, n] = signal_power / (interference_power + noise_power)


    return SINR

def normalize_columns(X):
    # 计算每一列的和
    col_sum = np.sum(X, axis=0)
    # 为避免除以 0，可以在除法前检查或加一个小常数 epsilon
    epsilon = 1e-10
    return X / (col_sum + epsilon)


def findMcsIndex(sinr):
    if math.isnan(sinr):
        return -1
    mcs_index = -1
    sinr_int = int(sinr) + 8
    if sinr_int < 0:
        mcs_index = -1
    elif sinr_int > 36:
        mcs_index = SINR_TABLE[36]
    else:
        mcs_index = SINR_TABLE[sinr_int]

    return mcs_index

def compute_MCS(X, H, V, Phi,noise_power=1e-8):
    """
    计算每个用户在各资源块上的SINR

    参数:
    -----------
    X : np.array, shape (M, N)
        用户-资源块分配矩阵（0或1），X[m,n]==1 表示用户 m 被分配了资源块 n。
    H : np.array, shape (M, N, Tx)
        信道矩阵，描述每个用户在各资源块、各发射天线上的信道响应（可以为复数）。
    V : np.array, shape (M, N, Tx)
        权重/预编码矩阵，与信道矩阵对应（可为复数）。
    noise_power : float, 可选
        噪声功率，默认为 1e-3。

    返回:
    --------
    SINR : np.array, shape (M, N)
        每个用户在对应资源块上的SINR值，对于未分配的资源块，SINR值保持为0。
    """
    M, N, Tx = H.shape
    SINR = np.zeros((M, N))
    MCS = np.zeros((M, N))-1


    for m in range(M):
        for n in range(N):
            # 若用户 m 在资源块 n 上被分配
            if X[m, n] != 0:
                # 期望信号：用户 m 在 n 资源块上的内积
                desired = np.dot(H[m, n, :], V[m, n, :])
                signal_power = X[m, n]*np.abs(desired) ** 2

                interference_power = 0.0
                # 对于同一资源块 n 上分配给其他用户的信号，计算干扰
                for k in range(M):
                    if k != m and X[k, n] != 0:
                        interfering = np.dot(H[m, n, :], V[k, n, :])
                        print("干扰:",H[m, n, :],V[k, n, :],interfering)
                        interference_power += X[k, n]*np.abs(interfering) ** 2
                SINR[m, n] = 10*np.log(signal_power / (interference_power + noise_power+abs(Phi[m,n])))
                MCS[m,n]=findMcsIndex(SINR[m, n])

    return SINR,MCS


def compute_sinr_loop(P,H, V, noise_power=1e-3,Phi=None):
    """
    计算不同用户在不同资源块上的 SINR（使用双重循环实现）

    输入:
      H: 信道矩阵，形状为 (N, M, Tx)
         N: 用户数
         M: 资源块数 (RBG数)
         Tx: 发射天线数
      V: 预编码矩阵，形状为 (N, Tx, M)
         对于每个资源块 m，V[n, :, m] 是对应用户 n 的预编码向量
      noise_power: 噪声功率 (默认1.0)

    输出:
      sinr: 每个用户在每个资源块上的SINR, 形状为 (N, M)
    """
    N, M, Tx = H.shape
    sinr = np.zeros((N, M))
    MCS = np.zeros((N, M)) - 1

    # 对于每个资源块 m 和每个用户 n
    for m in range(M):
        for n in range(N):
            # 计算期望信号功率：用户 n 自己的信道与预编码向量点乘
            if P[n,m]!=0:
                desired = np.dot(H[n, m, :], V[n, :, m])
                desired_power = P[n,m]*np.abs(desired) ** 2


                # 计算干扰功率：其他用户 k (k≠n) 的贡献
                interference_power = 0
                for k in range(N):
                    if k == n:
                        continue
                    interf = np.dot(H[n, m, :], V[k, :, m])
                    interference_power += P[k,m]*np.abs(interf) ** 2
                # x=desired_power / (interference_power + noise_power)

                sinr[n, m] = 10*np.log10(desired_power / (interference_power + noise_power))
                if Phi is not None:
                    sinr[n, m] = 10 * np.log10(desired_power / (interference_power + noise_power+abs(Phi[n,m])))
                    # print(interference_power,noise_power,abs(Phi[n,m]))

                MCS[n, m] = findMcsIndex(sinr[n, m])
            else:
                sinr[n, m] =-10000
                MCS[n, m] = -1

    return sinr,MCS

def getBaselineMcsSinr( MCS, Sinr):
    N,M=MCS.shape
    ave_mcs = np.zeros([ N])
    ave_sinr = np.zeros([ N])
    ave_count = np.zeros([ N])

    for k in range(N):
        mcs_count = 0
        total_mcs = 0
        total_sinr = 0
        for r in range(M):
            if MCS[k, r] == -1:
                continue
            else:
                total_mcs += MCS[k, r]
                total_sinr += Sinr[k, r]
                mcs_count += 1
        if mcs_count != 0:
            ave_mcs[k] = total_mcs // mcs_count
            ave_sinr[k] = total_sinr / mcs_count
        else:
            ave_mcs[k] = -1
            ave_sinr[k] = -10000
        ave_count[k] = mcs_count
    return ave_mcs, ave_count, ave_sinr


class Mcs_Sinr_p():
    def __init__(self,len_MCS=29,len_SINR=37):
        self.len_MCS = len_MCS
        self.len_SINR = len_SINR
        self.ori_P = np.zeros((len_SINR, len_MCS))
        self.SINR_class = []
        self.MCS_class = [i for i in range(29)]
        self.Function_list = []
    def load_table(self):
        workbook = load_workbook(filename='MCS_table.xlsx')
        # 选择工作表
        sheet = workbook['Sheet1']
        # 读取单元格内容
        cell_value = sheet['A1'].value
        # 读取整个工作表数据
        i = 0
        for row in sheet.iter_rows(min_row=1, max_col=30, values_only=True):
            if row[0] != 'SINR':
                self.SINR_class.append(row[0])
                self.ori_P[:][i] = row[1:]
                i += 1

    def create_inte(self):
        self.ori_P[np.isnan(self.ori_P)] = 0  # Y
        self.ori_P = np.array(self.ori_P)

        for j in range(len(self.MCS_class)):
            # 补0插值
            nums = self.len_SINR
            f = interpolate.interp1d(x=self.SINR_class[:nums], y=self.ori_P[:, j][:nums], kind="quadratic", bounds_error=False)
            self.Function_list.append(f)

    def Error_cal(self,pre_MCS,req_SINR,cls=29,probs=False):

        np.random.seed(42)
        UEs = pre_MCS.shape[0]
        error_p=[]
        for i in range(len(pre_MCS)):
            if pre_MCS[i] <= -1 or math.isnan( req_SINR[i]):
                p = 1
            else:
                if pre_MCS[i] >= 29:
                    pre_MCS[i] = 28
                p=self.Function_list[int(pre_MCS[i])](req_SINR[i])

            if np.isnan(p):
                # print("check!","MCS:",pre_MCS[i],"SINR:",req_SINR[i],"Probs:",p)
                #changed
                p=1
            elif p > 1:
                p = 1
            elif p < 0:
                p = 0

            error_p.append(p)
        event = (np.random.binomial(1, error_p, size=UEs))
        if probs==True:
            return error_p
        else:
            return event


    def Error_cal_single(self,pre_MCS,req_SINR,cls=29,probs=False):

        error_p=[]

        if req_SINR<=-1000:
            p=0
        if pre_MCS==-1:
            p=1
        else:
            p=self.Function_list[int(pre_MCS)](req_SINR)

        if np.isnan(p):
            # print("check!","MCS:",pre_MCS[i],"SINR:",req_SINR[i],"Probs:",p)
            #changed
            p=1
        if p<0:
            p=0
        if p>1:
            p=1
        error_p.append(p)
        event = (np.random.binomial(1, error_p[0], size=1))
        if probs==True:
            return error_p
        else:
            return event

def getSendRate( ave_mcs, ave_count):
    UserNum=len(ave_mcs)

    assert ave_mcs.shape == ave_count.shape, f"{ave_mcs.shape}, {ave_count.shape}"
    try:
        rate = np.zeros([UserNum])
        for k in range(UserNum):
            if ave_mcs[ k] != -1:
                rate[k] = int(144 * 16 * MCS_TABLE[ave_mcs[k].astype(int)] * ave_count[k])
    except:
        print(ave_mcs)
    return rate

def getCorrectedData( error_event, rate):
    return rate * (1 - error_event)


def logger_writer(instance_name,all_data,records,save_path=None):
    lst=0
    with open(f"./{instance_name}.pkl", "wb") as file:
        pickle.dump(lst, file)
    ##Prb

    ##UE satisfaction

    #event

    #MCS

    quit()
    pass

def generate_random_vector():
    # 生成三个随机数
    random_vector = np.random.rand(3)
    # 将向量归一化，使其和为1
    random_vector /= np.sum(random_vector)
    random_vector=np.array([1,1,1])
    return random_vector

def sample_index(tensor: torch.Tensor) -> tuple:
    torch.manual_seed(0)
    # 将tensor展平为一维
    flat_tensor = tensor.view(-1)
    # 对展平后的tensor计算softmax，得到概率分布
    probs = F.softmax(flat_tensor, dim=0)
    # 根据概率分布采样一个元素的索引（返回的是一个tensor）
    sampled_index = torch.multinomial(probs, num_samples=1)
    # 将采样到的索引转换为整数
    idx = sampled_index.item()
    # 将一维索引转换为二维索引（假设原tensor形状为 (n, m)）
    n, m = tensor.shape
    row = idx // m
    col = idx % m


    return (row, col)


def compute_B(A, k=300):
    """
    根据向量 A 和窗口大小 k 计算向量 B
    参数:
      A -- 数值列表或一维数组
      k -- 窗口大小（正整数）
    返回:
      B -- 生成的新向量，满足:
            B[0] = A[0]
            B[1] = (A[0] + A[1]) / 2
            ...
            B[k-1] = (A[0] + A[1] + ... + A[k-1]) / k
            对于 i >= k, B[i] = (A[i-k+1] + ... + A[i]) / k
    """
    n = len(A)
    B = []
    for i in range(n):
        if i < k:
            # i 从 0 开始，所以前 i+1 个元素
            window = A[:i+1]
            avg = sum(window) / (i+1)
        else:
            # 对于 i >= k，取最近 k 个元素：A[i-k+1] 到 A[i]
            window = A[i-k+1:i+1]
            avg = sum(window) / k
        B.append(avg)
    return B